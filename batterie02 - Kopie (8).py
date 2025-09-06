# Batterie-Bewirtschaftung. Anhand von Ausgangsparametern wird versucht eine Batterie wirtschatlich optimal zu Laden und zu Entladen.
# Parameter: batterie_kapazitaet, lade_menge, ladung_min, ladung_max, speicher_verluste, entlade_verluste
# INPUT: eine Exceltabelle mit Zeitstempel, Preis und Verbrauch für ein Jahr auf 15 Minuten Basis
# Output: Eine Exceltabelle mit den optimalen Lade- und Entlademenge auf 15 Minuten Basis
# 2026 September, 

import openpyxl
import dateparser
import matplotlib.pyplot as plt
import pulp
import requests
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Font

def lade_daten(excel_datei):
    wb = openpyxl.load_workbook(excel_datei, data_only=True)
    ws = wb.active
    daten = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= 1000:
            break
        datum = row[0]
        try:
            preis = float(str(row[1]).replace(',', '.'))
        except (ValueError, TypeError):
            preis = 0
        try:
            verbrauch = float(str(row[5]).replace(',', '.'))
        except (ValueError, TypeError):
            verbrauch = 0
        daten.append([datum, preis, verbrauch])
    return daten

def filter_zeitraum(daten, von, bis):
    von_dt = dateparser.parse(von)
    bis_dt = dateparser.parse(bis)
    return [row for row in daten if dateparser.parse(str(row[0])) and von_dt <= dateparser.parse(str(row[0])) <= bis_dt]

def batterie_bewirtschaftung(daten, batterie_kapazitaet=10.0, lade_menge=0.5, ladung_min=10, ladung_max=90, speicher_verluste=10, entlade_verluste=10):
    speicher_verluste = speicher_verluste / 100.0                     # Prozent zu Anteil
    entlade_verluste = entlade_verluste / 100.0
    ladung_min_anteil = ladung_min / 100.0                            # Umrechnung Prozent auf Anteil
    ladung_max_anteil = ladung_max / 100.0
    n = len(daten)

    # Variablen
    l = [pulp.LpVariable(f"l_{t}", lowBound=0, upBound=lade_menge) for t in range(n)]             
    e = [pulp.LpVariable(f"e_{t}", lowBound=0, upBound=daten[t][2]) for t in range(n)]   # e[t] ist die tatsächlich genutzte Energie (Verbrauch)
    s = [pulp.LpVariable(f"s_{t}", lowBound=ladung_min_anteil * batterie_kapazitaet, upBound=ladung_max_anteil * batterie_kapazitaet) for t in range(n+1)]

    prob = pulp.LpProblem("BatterieOptimierung", pulp.LpMaximize)

    prob += pulp.lpSum([                                               # Zielfunktion: Erlöse - Kosten (mit Verlusten)
        daten[t][1] * e[t] - daten[t][1] * l[t] / (1 - speicher_verluste)
        for t in range(n)])
    
    prob += s[0] == ladung_min_anteil * batterie_kapazitaet            # Anfangs- und Endfüllstand
    prob += s[n] == ladung_min_anteil * batterie_kapazitaet

    for t in range(n):                                                 # Speicherentwicklung
        prob += s[t+1] == s[t] + l[t] - e[t] / (1 - entlade_verluste)  # Für den Verbrauch e[t] muss die Batterie um e[t] / (1 - entlade_verluste) entladen werden

    last_tag = None                                                    # Batterie am Tagesende auf ladung_min
    for t in range(n):
        datum = str(daten[t][0])
        tag = datum[:10]
        if last_tag and tag != last_tag:
            prob += s[t] == ladung_min_anteil * batterie_kapazitaet
        last_tag = tag

    prob.solve(pulp.PULP_CBC_CMD(msg=0))                               # Lösen

    aktionen = []                                                      # Ergebnisse aufbereiten
    for t in range(n):
        datum, preis, verbrauch = daten[t]
        lade_val = round(l[t].value() or 0, 4)                         # Menge, die in die Batterie kommt
        verbrauch_val = round(e[t].value() or 0, 4)                    # Tatsächlich genutzte Energie
        entlade_val = round(verbrauch_val / (1 - entlade_verluste), 4) if verbrauch_val > 0 else 0.0  # Entnommene Energie aus Batterie
        speicher_val = round(s[t+1].value() or 0, 4)
        menge_i = round(verbrauch, 4)
        preis = round(preis, 4)
        netz_laden = round(lade_val / (1 - speicher_verluste), 4) if lade_val > 0 else 0.0
        if lade_val > 0:
            betrag = round(netz_laden * preis, 4)
            aktionen.append((datum, "Laden", preis, speicher_val, None, betrag, menge_i, netz_laden))
        elif verbrauch_val > 0:
            betrag = round(-verbrauch_val * preis, 4)
            aktionen.append((datum, "Entladen", preis, speicher_val, None, betrag, menge_i, entlade_val))
        else:
            aktionen.append((datum, "Nichts", preis, speicher_val, None, 0.0, menge_i, 0.0))
    return aktionen

def schreibe_excel(aktionen, dateiname, parameter):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batterie Bewirtschaftung"

    # Parameter-Header (3 Zeilen)
    ws.cell(row=1, column=1, value=f"Von-Datum: {parameter.get('von_datum', '')}")
    ws.cell(row=1, column=3, value=f"Min. Ladezustand in % : {parameter.get('ladung_min', '')}")
    ws.cell(row=1, column=5, value=f"Ladeverluste    in %  : {parameter.get('speicher_verluste', '')}")

    ws.cell(row=2, column=1, value=f"Bis-Datum : {parameter.get('bis_datum', '')}")
    ws.cell(row=2, column=3, value=f"Max. Ladezustand in %: {parameter.get('ladung_max', '')}")
    ws.cell(row=2, column=5, value=f"Entladeverluste in %: {parameter.get('entlade_verluste', '')}")

    ws.cell(row=3, column=1, value=f"Batteriekapazität in kWh: {parameter.get('batterie_kapazitaet', '')}")
    ws.cell(row=3, column=3, value=f"Max. Lade-/Entlademenge kWh: {parameter.get('lade_menge', '')}")

    # Leerzeile
    start_row = 5

    header = ['Datum', 'Aktion', 'Menge_Verbrauch', 'Menge_Laden', 'Menge_Entladen', 'Preis', 'Fuellstand', 'Betrag_Laden', 'Betrag_Entladung']
    for col_num, col_name in enumerate(header, 1):
        cell = ws.cell(row=start_row, column=col_num, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].width = 17

    # Summenbildung
    tagessummen = {}
    monatssummen = {}
    gesamtsummen = [0.0] * (len(header) - 2)
    for zeile in aktionen:
        datum = zeile[0]
        if datum is None:
            continue
        tag = dateparser.parse(str(datum)).date().isoformat()
        monat = tag[:7]
        # Summenfelder gemäß Überschrift (ohne Preis und Füllstand)
        menge_verbrauch = float(zeile[6]) if zeile[6] not in ('', None) else 0.0
        menge_laden = float(zeile[7]) if zeile[1] == "Laden" and zeile[7] not in ('', None) else 0.0
        menge_entladen = float(zeile[7]) if zeile[1] == "Entladen" and zeile[7] not in ('', None) else 0.0
        betrag_laden = float(zeile[5]) if zeile[1] == "Laden" and zeile[5] not in ('', None) else 0.0
        betrag_entladung = float(zeile[5]) if zeile[1] == "Entladen" and zeile[5] not in ('', None) else 0.0
        # Summen: [Menge_Verbrauch, Menge_Laden, Menge_Entladen, Betrag_Laden, Betrag_Entladung]
        summenwerte = [menge_verbrauch, menge_laden, menge_entladen, betrag_laden, betrag_entladung]
        # Tagessummen
        if tag not in tagessummen:
            tagessummen[tag] = [0.0] * len(summenwerte)
        tagessummen[tag] = [a + b for a, b in zip(tagessummen[tag], summenwerte)]
        # Monatssummen
        if monat not in monatssummen:
            monatssummen[monat] = [0.0] * len(summenwerte)
        monatssummen[monat] = [a + b for a, b in zip(monatssummen[monat], summenwerte)]
        # Gesamtsummen
        gesamtsummen = [a + b for a, b in zip(gesamtsummen, summenwerte)]

    # Zeilen schreiben und Tagessummen einfügen
    for i, zeile in enumerate(aktionen):
        datum = zeile[0]
        if datum is None:
            continue
        tag = dateparser.parse(str(datum)).date().isoformat()
        menge_laden = zeile[7] if zeile[1] == "Laden" else 0.0
        menge_entladen = zeile[7] if zeile[1] == "Entladen" else 0.0
        betrag_laden = zeile[5] if zeile[1] == "Laden" else 0.0
        betrag_entladung = zeile[5] if zeile[1] == "Entladen" else 0.0
        neue_zeile = [
            zeile[0],           # Datum
            zeile[1],           # Aktion
            zeile[6],           # Menge_Verbrauch
            menge_laden,        # Menge_Laden
            menge_entladen,     # Menge_Entladen
            zeile[2],           # Preis
            zeile[3],           # Füllstand
            betrag_laden,       # Betrag_Laden
            betrag_entladung    # Betrag_Entladung
        ]
        # Formatierung: Spalte A zentrieren, Spalte C bis I: >0 mit 4 Nachkommastellen, 0 als leer
        formatierte_zeile = [
            neue_zeile[0],
            neue_zeile[1],
        ] + [float(x.replace(',', '.')) if isinstance(x, (int, float)) or (isinstance(x, str) and x.replace(',', '').replace('.', '').isdigit()) and x != '' else '' for x in [f'{v:.4f}'.replace('.', ',') if isinstance(v, (int, float)) and v != 0 else '' for v in neue_zeile[2:]]]
        ws.append(formatierte_zeile)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        for col in range(2, 10):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    # Tagessummen schreiben
        next_tag = None
        if i + 1 < len(aktionen):
            next_datum = aktionen[i+1][0]
            if next_datum is not None:
                next_tag = dateparser.parse(str(next_datum)).date().isoformat()
        if (i + 1 == len(aktionen)) or (next_tag is not None and next_tag != tag):
            sum_row = [tag, 'Tagessumme'] + [float(x.replace(',', '.')) if x != '' else '' for x in [f'{v:.4f}'.replace('.', ',') if v != 0 else '' for v in tagessummen[tag]]]
            sum_row = sum_row[:5] + ['', ''] + sum_row[5:]
            ws.append(sum_row)
            for col in range(1, len(header)+1):
                ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
                ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    # Monatssummen
    for monat, summen in monatssummen.items():
        sum_row = [monat, 'Monatssumme'] + [float(x.replace(',', '.')) if x != '' else '' for x in [f'{v:.4f}'.replace('.', ',') if v != 0 else '' for v in summen]]
        sum_row = sum_row[:5] + ['', ''] + sum_row[5:]
        ws.append(sum_row)
        for col in range(1, len(header)+1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    # Gesamtsumme
    sum_row = ['Gesamt', 'Gesamtsumme'] + [float(x.replace(',', '.')) if x != '' else '' for x in [f'{v:.4f}'.replace('.', ',') if v != 0 else '' for v in gesamtsummen]]
    sum_row = sum_row[:5] + ['', ''] + sum_row[5:]
    ws.append(sum_row)
    for col in range(1, len(header)+1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    wb.save(dateiname)

def visualisiere(aktionen):
    lade_zeitpunkte = [dateparser.parse(str(a[0])) for a in aktionen if a[1] == 'Laden']
    lade_preise_plot = [a[2] for a in aktionen if a[1] == 'Laden']
    entlade_zeitpunkte = [dateparser.parse(str(a[0])) for a in aktionen if a[1] == 'Entladen']
    entlade_preise_plot = [a[2] for a in aktionen if a[1] == 'Entladen']
    plt.figure(figsize=(14, 6))
    plt.plot(lade_zeitpunkte, lade_preise_plot, 'go', label='Laden')
    plt.plot(entlade_zeitpunkte, entlade_preise_plot, 'ro', label='Entladen')
    plt.xlabel('Zeit')
    plt.ylabel('Preis (Cent/kWh)')
    plt.title('Batterie-Lade- und Entladezeitpunkte')
    plt.legend()
    plt.grid(True)

def marktpreise_von_awatar():
    url = "https://api.awattar.at/v1/marketdata"
    try:
        response = requests.get(url)
        print("Awattar API Response Status:", response.status_code)
        if response.status_code == 200:
            data = response.json().get("data", [])
            print("Awattar Zeitfenster:")
            for entry in data:
                start_dt = datetime.fromtimestamp(entry["start_timestamp"] / 1000)
                end_dt = datetime.fromtimestamp(entry["end_timestamp"] / 1000)
                print(f"{start_dt.strftime('%Y-%m-%d %H:%M')} - {end_dt.strftime('%Y-%m-%d %H:%M')}: {entry['marketprice']} {entry['unit']}")
        else:
            print("Awattar API Response:")
            print(response.text[:1000])
    except Exception as e:
        print("Fehler beim Abrufen der Marktpreise:", e)

def marktpreise_von_entsoe():
    token = "1d8d5cf8-299c-4b8f-b142-3f3b294f2b77"
    morgen = datetime.now() + timedelta(days=1)
    start = morgen.strftime("%Y%m%d0000")
    end = morgen.strftime("%Y%m%d2300")

    #https://web-api.transparency.entsoe.eu/api?securityToken=1d8d5cf8-299c-4b8f-b142-3f3b294f2b77&documentType=A44&in_Domain=10YAT-APG------L&out_Domain=10YAT-APG------L&periodStart=202509020000&periodEnd=202509022300

    url = (
    f"https://web-api.transparency.entsoe.eu/api?securityToken={token}"
    f"&documentType=A44"
    f"&in_Domain=10YAT-APG------L"
    f"&out_Domain=10YAT-APG------L"
    f"&periodStart={start}"
    f"&periodEnd={end}"
)
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        response = requests.get(url, headers=headers)
        print("ENTSO-E API Response Status:", response.status_code)
        if response.status_code == 200:
            import xml.etree.ElementTree as ET
            root = ET.fromstring(response.content)
            ns = {'ns': 'urn:iec62325.351:tc57wg16:451-3:publicationdocument:7:0'}
            prices = []
            for time_series in root.findall('.//ns:TimeSeries', ns):
                for period in time_series.findall('.//ns:Period', ns):
                    start_time = period.find('ns:timeInterval/ns:start', ns).text
                    for point in period.findall('ns:Point', ns):
                        position = int(point.find('ns:position', ns).text)
                        price = float(point.find('ns:price.amount', ns).text)
                        dt = datetime.strptime(start_time, "%Y-%m-%dT%H:%MZ") + timedelta(hours=position-1)
                        prices.append((dt, price))
            print("ENTSO-E Day-Ahead Preise für Österreich (EUR/MWh):")
            for dt, price in prices:
                print(f"{dt.strftime('%Y-%m-%d %H:%M')}: {price:.2f} EUR/MWh")
        else:
            print("ENTSO-E API Response:")
            print(response.text[:1000])
    except Exception as e:
        print("Fehler beim Abrufen der ENTSO-E Marktpreise:", e)

def main():
    excel_datei_input = "H0_Preise_15minuten.xlsx"  # Hier ggf. den Dateinamen anpassen

    # Default-Werte für die Eingabeparameter
    von_default = "2026-01-01"
    bis_default = "2026-01-03"
    batterie_kapazitaet_default = 10.0
    lade_menge_default = 0.5
    ladung_min_default = 10.0
    ladung_max_default = 90.0
    speicher_verluste_default = 5.0
    entlade_verluste_default = 5.0
    
    von = input(f"Von-Datum (YYYY-MM-DD)    [Default: {von_default} ]: ")
    bis = input(f"Bis-Datum (YYYY-MM-DD)    [Default: {bis_default} ]: ")
    batterie_kapazitaet = input(f"Batteriekapazität in kWh        [Default: {batterie_kapazitaet_default} ]: ")
    lade_menge = input(f"Maximale Lade-/Entlademenge kWh [Default:  {lade_menge_default} ]: ")
    ladung_min = input(f"Minimaler Ladezustand in %      [Default: {ladung_min_default} ]: ")
    ladung_max = input(f"Maximaler Ladezustand in %      [Default: {ladung_max_default} ]: ")
    speicher_verluste = input(f"Ladeverluste    in %            [Default:  {speicher_verluste_default} ]: ")
    entlade_verluste = input(f"Entladeverluste in %            [Default:  {entlade_verluste_default} ]: ")

    if not von: von = von_default
    if not bis: bis = bis_default
    batterie_kapazitaet = float(batterie_kapazitaet) if batterie_kapazitaet else batterie_kapazitaet_default
    lade_menge = float(lade_menge) if lade_menge else lade_menge_default
    ladung_min = float(ladung_min) if ladung_min else ladung_min_default
    ladung_max = float(ladung_max) if ladung_max else ladung_max_default
    speicher_verluste = float(speicher_verluste) if speicher_verluste else speicher_verluste_default
    entlade_verluste = float(entlade_verluste) if entlade_verluste else entlade_verluste_default

    von_dt = dateparser.parse(von + " 00:00")         # Setze Zeit auf Tagesanfang/ende
    bis_dt = dateparser.parse(bis + " 23:45")

    daten = lade_daten(excel_datei_input)
    daten = [row for row in daten if dateparser.parse(str(row[0])) and von_dt <= dateparser.parse(str(row[0])) <= bis_dt]

    aktionen = batterie_bewirtschaftung(daten, batterie_kapazitaet, lade_menge, ladung_min, ladung_max, speicher_verluste, entlade_verluste)
    # Parameter für Excel-Header
    parameter = {
        'von_datum': von,
        'bis_datum': bis,
        'batterie_kapazitaet': batterie_kapazitaet,
        'lade_menge': lade_menge,
        'ladung_min': ladung_min,
        'ladung_max': ladung_max,
        'speicher_verluste': speicher_verluste,
        'entlade_verluste': entlade_verluste
    }
    # Übergabe der Parameter direkt an schreibe_excel
    schreibe_excel(aktionen, 'batterie_aktionen.xlsx', parameter)
    visualisiere(aktionen)

if __name__ == "__main__":
    #marktpreise__von_awatar()
    #marktpreise_von_entsoe()
    main()
