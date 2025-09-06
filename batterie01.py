# Aufbereiten der Inputdatei - Stundenpreise in 15 Minuten Preise umwandeln.

import openpyxl
import os

wb = openpyxl.load_workbook('H0_Preise.xlsx')  # 
ws = wb.active

start_row = 4
max_row = ws.max_row

# Für jede Originalzeile ab Zeile 4: verschiebe 3 Zellen nach unten und trage die neuen Werte ein
for idx in range(max_row - 3):
    row = start_row + idx * 4
    if idx % 100 == 0 and idx > 0:
        print(f"Fortschritt: {idx} Schleifendurchläufe")
    
    # Verschiebe 3 Zellen nach unten in Spalte A und B
    for col in ['A', 'B']:
        for move_row in range(max_row + 3, row, -1):
            ws[f'{col}{move_row}'] = ws[f'{col}{move_row-3}'].value
    
    # Fülle die 3 neuen Zellen in Spalte A mit Wert aus Spalte D der gleichen Zeilennummer
    for i in range(1, 4):
        ws[f'A{row + i}'] = ws[f'D{row + i}'].value
        ws[f'B{row + i}'] = ws[f'B{row}'].value

ziel_datei = 'H0_Preise_quartalsweise.xlsx'
if os.path.exists(ziel_datei):
    os.remove(ziel_datei)
wb.save(ziel_datei)